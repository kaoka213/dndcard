import json
import os
import re
import time
import unicodedata
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup

XLSX_PATH = r"c:\Projeler\DNDCard\Prompt_Kaynaklari_V2.xlsx"
OUTPUT_ROOT = r"c:\Projeler\DNDCard\prompts"
SHEET_NAME = "🗄️ Tüm Kaynaklar (Veritabanı)"
ENV_PATH = r"c:\Projeler\DNDCard\.env"
ROW_START = 2
ROW_END = 35
OUTPUT_FILE = "agents.json"


def get_github_token() -> str:
	value = os.getenv("GITHUB_TOKEN", "").strip()
	if value:
		return value
	if not os.path.exists(ENV_PATH):
		return ""
	with open(ENV_PATH, "r", encoding="utf-8") as env_file:
		for line in env_file:
			line = line.strip()
			if not line or line.startswith("#") or "=" not in line:
				continue
			key, raw = line.split("=", 1)
			if key.strip() == "GITHUB_TOKEN":
				return raw.strip().strip('"').strip("'")
	return ""


def slugify(text: str) -> str:
	text = text.strip().lower()
	text = unicodedata.normalize("NFKD", text)
	text = text.encode("ascii", "ignore").decode("ascii")
	text = re.sub(r"[^a-z0-9]+", "-", text)
	text = text.strip("-")
	return text or "uncategorized"


def get_cell_text(cell) -> str:
	return "" if cell.value is None else str(cell.value).strip()


def get_hyperlink(cell) -> str:
	if cell.hyperlink and cell.hyperlink.target:
		return str(cell.hyperlink.target).strip()
	value = get_cell_text(cell)
	if value.startswith("http://") or value.startswith("https://"):
		return value
	return ""


def extract_links(ws, header_row_idx: int, col_map: dict) -> list[dict]:
	records = []
	min_row = max(header_row_idx + 1, ROW_START) if ROW_START else header_row_idx + 1
	max_row = ROW_END if ROW_END else None
	for row in ws.iter_rows(min_row=min_row, max_row=max_row):
		no_val = get_cell_text(row[col_map["no"]])
		if not no_val:
			continue
		link_cell = row[col_map["link"]]
		url = get_hyperlink(link_cell)
		if not url:
			# Fallback: search any cell for a URL
			for cell in row:
				val = get_cell_text(cell)
				if val.startswith("http://") or val.startswith("https://"):
					url = val
					break
		record = {
			"no": no_val,
			"source_type": get_cell_text(row[col_map["source_type"]]),
			"focus_area": get_cell_text(row[col_map["focus_area"]]),
			"project_name": get_cell_text(row[col_map["project_name"]]),
			"owner": get_cell_text(row[col_map["owner"]]),
			"level": get_cell_text(row[col_map["level"]]),
			"use_case": get_cell_text(row[col_map["use_case"]]),
			"sheet_description": get_cell_text(row[col_map["sheet_description"]]),
			"url": url,
		}
		records.append(record)
	return records


def find_header_row(ws) -> tuple[int, dict]:
	header_row_idx = None
	header_values = None
	for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
		values = [get_cell_text(c) for c in row]
		if "Hızlı Bağlantı" in values and "Odak Alanı" in values:
			header_row_idx = idx
			header_values = values
			break
	if header_row_idx is None:
		raise ValueError("Header row not found")

	def col_index(name: str) -> int:
		return header_values.index(name)

	col_map = {
		"no": col_index("No"),
		"source_type": col_index("Kaynak Türü"),
		"focus_area": col_index("Odak Alanı"),
		"project_name": col_index("Proje / Araç Adı"),
		"owner": col_index("Geliştirici / Sahibi"),
		"level": col_index("Zorluk & Seviye"),
		"use_case": col_index("En İyi Kullanım Senaryosu (Ne için kullanılır?)"),
		"sheet_description": col_index("Detaylı Açıklama"),
		"link": col_index("Hızlı Bağlantı"),
	}
	return header_row_idx, col_map


def safe_request(url: str) -> str:
	headers = {
		"User-Agent": "Mozilla/5.0 (PromptScraper/1.0; +https://example.com)"
	}
	resp = requests.get(url, headers=headers, timeout=30)
	resp.raise_for_status()
	return resp.text


def safe_request_json(url: str) -> dict:
	headers = {
		"User-Agent": "Mozilla/5.0 (PromptScraper/1.0; +https://example.com)",
		"Accept": "application/vnd.github+json",
	}
	github_token = get_github_token()
	if github_token:
		headers["Authorization"] = f"Bearer {github_token}"
	resp = requests.get(url, headers=headers, timeout=30)
	resp.raise_for_status()
	return resp.json()


def parse_github_repo(url: str) -> tuple[str, str] | tuple[None, None]:
	parsed = urlparse(url)
	if "github.com" not in parsed.netloc:
		return None, None
	parts = parsed.path.strip("/").split("/")
	if len(parts) < 2:
		return None, None
	return parts[0], parts[1]


def extract_prompt_blocks(soup: BeautifulSoup) -> list[str]:
	prompts = []
	for tag in soup.find_all(["pre", "code"]):
		text = tag.get_text("\n", strip=True)
		if len(text) >= 20:
			prompts.append(text)

	for tag in soup.find_all(True, {"class": re.compile("prompt|instruction|example", re.I)}):
		text = tag.get_text("\n", strip=True)
		if len(text) >= 20:
			prompts.append(text)

	for tag in soup.find_all(True, {"id": re.compile("prompt|instruction|example", re.I)}):
		text = tag.get_text("\n", strip=True)
		if len(text) >= 20:
			prompts.append(text)

	# Deduplicate while preserving order
	seen = set()
	deduped = []
	for text in prompts:
		key = re.sub(r"\s+", " ", text).strip()
		if key and key not in seen:
			seen.add(key)
			deduped.append(text)
	return deduped[:20]


def extract_meta(soup: BeautifulSoup, base_url: str) -> tuple[str, str]:
	description = ""
	icon_url = ""

	og_desc = soup.find("meta", {"property": "og:description"})
	if og_desc and og_desc.get("content"):
		description = og_desc["content"].strip()
	if not description:
		meta_desc = soup.find("meta", {"name": "description"})
		if meta_desc and meta_desc.get("content"):
			description = meta_desc["content"].strip()
	if not description:
		first_p = soup.find("p")
		if first_p:
			description = first_p.get_text(" ", strip=True)

	og_img = soup.find("meta", {"property": "og:image"})
	if og_img and og_img.get("content"):
		icon_url = og_img["content"].strip()
	if not icon_url:
		for link_tag in soup.find_all("link"):
			rels = link_tag.get("rel") or []
			rels = [r.lower() for r in rels] if isinstance(rels, list) else [str(rels).lower()]
			if any(r in rels for r in ["icon", "shortcut icon", "apple-touch-icon"]):
				if link_tag.get("href"):
					icon_url = link_tag["href"].strip()
					break

	if icon_url and not icon_url.startswith("http"):
		icon_url = urljoin(base_url, icon_url)
	return description, icon_url


def extract_agent_fields_from_page(soup: BeautifulSoup, base_url: str) -> dict:
	name = ""
	description = ""
	icon_url = ""
	prompt_text = ""
	tags: list[str] = []

	if soup.title and soup.title.get_text(strip=True):
		name = soup.title.get_text(strip=True)

	h1 = soup.find("h1")
	if h1:
		name = h1.get_text(" ", strip=True) or name
		# Description often appears in the next paragraph after the title.
		next_p = h1.find_next("p")
		if next_p:
			description = next_p.get_text(" ", strip=True)

	# Tag chips/badges near the header.
	for tag in soup.find_all(True, {"class": re.compile(r"badge|tag|chip", re.I)}):
		text = tag.get_text(" ", strip=True)
		if 1 <= len(text) <= 30:
			tags.append(text)

	# System Prompt section
	for header in soup.find_all(["h2", "h3", "h4", "strong"]):
		if header.get_text(" ", strip=True).lower() == "system prompt":
			cursor = header
			for _ in range(8):
				cursor = cursor.find_next()
				if cursor is None:
					break
				if cursor.name in ["pre", "code", "textarea"]:
					text = cursor.get_text("\n", strip=True)
					if len(text) >= 20:
						prompt_text = text
						break
			if prompt_text:
				break

	# Fallbacks
	if not description:
		description, icon_url = extract_meta(soup, base_url)
	if not icon_url:
		_, icon_url = extract_meta(soup, base_url)

	# Deduplicate tags
	clean_tags = []
	seen = set()
	for item in tags:
		key = item.lower()
		if key in seen:
			continue
			
		seen.add(key)
		clean_tags.append(item)

	return {
		"agent_name": name.strip(),
		"agent_description": description.strip(),
		"agent_icon": icon_url.strip(),
		"agent_tags": clean_tags,
		"agent_prompt": prompt_text.strip(),
	}


def extract_from_github(url: str) -> dict:
	owner, repo = parse_github_repo(url)
	if not owner or not repo:
		return {}
	candidates = [
		f"https://raw.githubusercontent.com/{owner}/{repo}/main/README.md",
		f"https://raw.githubusercontent.com/{owner}/{repo}/master/README.md",
	]
	for raw_url in candidates:
		try:
			text = safe_request(raw_url)
		except Exception:
			continue
		description = ""
		for line in text.splitlines():
			line = line.strip()
			if line and not line.startswith("#"):
				description = line
				break
		prompt_blocks = []
		code_block = []
		in_code = False
		for line in text.splitlines():
			if line.startswith("```"):
				if in_code and code_block:
					prompt_blocks.append("\n".join(code_block))
					code_block = []
				in_code = not in_code
				continue
			if in_code:
				code_block.append(line)
		return {
			"prompt_description": description,
			"prompts": prompt_blocks[:20],
		}
	return {}


def extract_github_meta(url: str) -> dict:
	owner, repo = parse_github_repo(url)
	if not owner or not repo:
		return {}
	try:
		repo_info = safe_request_json(f"https://api.github.com/repos/{owner}/{repo}")
	except Exception:
		return {}
	name = repo_info.get("name", "")
	description = repo_info.get("description", "") or ""
	tags = repo_info.get("topics", []) or []
	icon = repo_info.get("owner", {}).get("avatar_url", "")
	return {
		"agent_name": str(name).strip(),
		"agent_description": str(description).strip(),
		"agent_tags": tags,
		"agent_icon": str(icon).strip(),
	}


def is_agent_like(text: str) -> bool:
	value = re.sub(r"\s+", " ", text).strip().lower()
	if len(value) < 30 or len(value) > 5000:
		return False
	keywords = ["you are", "system prompt", "system:", "role", "agent", "instructions", "persona"]
	if not any(k in value for k in keywords):
		return False
	codey = re.search(r"\b(import|export|function|class|const|let|var|<html|<div|@tailwind|react)\b", value)
	if codey and "you are" not in value and "system prompt" not in value:
		return False
	return True


def extract_agent_texts_from_content(content: str) -> list[str]:
	agents = []

	# Markdown code fences
	block = []
	in_code = False
	for line in content.splitlines():
		if line.strip().startswith("```"):
			if in_code and block:
				text = "\n".join(block).strip()
				if is_agent_like(text):
					agents.append(text)
				block = []
			in_code = not in_code
			continue
		if in_code:
			block.append(line)

	# Simple key-based extraction for yaml-like text
	key_pattern = re.compile(r"^(system|prompt|instructions?|role|persona)\s*:\s*(.+)$", re.I)
	for line in content.splitlines():
		match = key_pattern.match(line.strip())
		if match:
			text = match.group(2).strip()
			if is_agent_like(text):
				agents.append(text)

	# Filter full-text blocks
	for chunk in content.split("\n\n"):
		text = chunk.strip()
		if is_agent_like(text):
			agents.append(text)

	# Deduplicate
	cleaned = []
	seen = set()
	for text in agents:
		key = re.sub(r"\s+", " ", text).strip()
		if key in seen:
			continue
		seen.add(key)
		cleaned.append(text)
	return cleaned[:50]


def extract_agent_texts_from_json(data) -> list[str]:
	results = []
	if isinstance(data, dict):
		for key, value in data.items():
			if isinstance(value, (dict, list)):
				results.extend(extract_agent_texts_from_json(value))
			elif isinstance(value, str):
				if re.search(r"^(system|prompt|instructions?|role|persona)$", str(key), re.I):
					if is_agent_like(value):
						results.append(value)
	elif isinstance(data, list):
		for item in data:
			results.extend(extract_agent_texts_from_json(item))
	return results


def extract_agents_from_github(url: str) -> list[dict]:
	owner, repo = parse_github_repo(url)
	if not owner or not repo:
		return []

	try:
		repo_info = safe_request_json(f"https://api.github.com/repos/{owner}/{repo}")
		branch = repo_info.get("default_branch", "main")
		tree = safe_request_json(
			f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch}?recursive=1"
		)
		files = tree.get("tree", [])
	except Exception:
		return []

	agent_files = []
	for entry in files:
		if entry.get("type") != "blob":
			continue
		path = entry.get("path", "")
		if not re.search(r"agent|prompt|system|instruction|persona|role", path, re.I):
			continue
		if not re.search(r"\.(md|txt|json|yaml|yml)$", path, re.I):
			continue
		agent_files.append(path)
		if len(agent_files) >= 50:
			break

	agents = []
	for path in agent_files:
		raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{path}"
		try:
			content = safe_request(raw_url)
		except Exception:
			continue
		if len(content) > 300_000:
			continue
		texts = []
		if path.lower().endswith(".json"):
			try:
				data = json.loads(content)
				texts = extract_agent_texts_from_json(data)
			except Exception:
				texts = extract_agent_texts_from_content(content)
		else:
			texts = extract_agent_texts_from_content(content)
		if texts:
			agents.append({
				"file": path,
				"count": len(texts),
				"texts": texts,
			})
		else:
			trimmed = content.strip()
			if len(trimmed) >= 30:
				max_len = 5000
				truncated = len(trimmed) > max_len
				agents.append({
					"file": path,
					"count": 1,
					"texts": [trimmed[:max_len]],
					"raw": True,
					"truncated": truncated,
				})
	return agents


def scrape_url(url: str) -> dict:
	result = {
		"prompt_description": "",
		"icon_url": "",
		"prompts": [],
		"agents": [],
		"agent_name": "",
		"agent_description": "",
		"agent_icon": "",
		"agent_tags": [],
		"agent_prompt": "",
		"error": "",
	}
	try:
		if "github.com" in url:
			gh_data = extract_from_github(url)
			result.update({k: v for k, v in gh_data.items() if v})
			result["agents"] = extract_agents_from_github(url)
			gh_meta = extract_github_meta(url)
			for key in ["agent_name", "agent_description", "agent_tags", "agent_icon"]:
				if gh_meta.get(key) and not result.get(key):
					result[key] = gh_meta[key]

		html = safe_request(url)
		soup = BeautifulSoup(html, "lxml")
		agent_fields = extract_agent_fields_from_page(soup, url)
		description, icon_url = extract_meta(soup, url)
		prompts = extract_prompt_blocks(soup)

		if not result["prompt_description"]:
			result["prompt_description"] = description
		if not result["icon_url"]:
			result["icon_url"] = icon_url
		if not result["prompts"]:
			result["prompts"] = prompts
		if not result["agents"]:
			result["agents"] = [
				{"file": "page", "count": len(prompts), "texts": prompts}
			]
		for key in ["agent_name", "agent_description", "agent_icon", "agent_tags", "agent_prompt"]:
			if agent_fields.get(key):
				result[key] = agent_fields[key]

		if not result.get("agent_prompt"):
			if result.get("agents"):
				first_agent = result["agents"][0]
				texts = first_agent.get("texts", []) if isinstance(first_agent, dict) else []
				if texts:
					result["agent_prompt"] = texts[0]
			if not result.get("agent_prompt") and result.get("prompts"):
				result["agent_prompt"] = result["prompts"][0]
	except Exception as exc:
		result["error"] = str(exc)
	return result


def main() -> None:
	wb = openpyxl.load_workbook(XLSX_PATH)
	if SHEET_NAME not in wb.sheetnames:
		raise ValueError(f"Sheet not found: {SHEET_NAME}")
	ws = wb[SHEET_NAME]

	header_row_idx, col_map = find_header_row(ws)
	records = extract_links(ws, header_row_idx, col_map)

	categorized: dict[str, list[dict]] = {}
	log_path = os.path.join(OUTPUT_ROOT, "_scrape_log.txt")
	if os.path.exists(log_path):
		os.remove(log_path)

	total = len(records)
	for idx, record in enumerate(records, start=1):
		url = record.get("url")
		if not url:
			record.update({"prompt_description": "", "icon_url": "", "prompts": [], "agents": []})
		else:
			scrape = scrape_url(url)
			record.update(scrape)
		category = slugify(record.get("focus_area", ""))
		categorized.setdefault(category, []).append(record)
		with open(log_path, "a", encoding="utf-8") as log_file:
			log_file.write(f"[{idx}/{total}] {url} | error={record.get('error','')}\n")
		print(f"{idx}/{total} {url}")
		time.sleep(0.2)

	for category, items in categorized.items():
		out_dir = os.path.join(OUTPUT_ROOT, category)
		os.makedirs(out_dir, exist_ok=True)
		out_path = os.path.join(out_dir, "items.json")
		with open(out_path, "w", encoding="utf-8") as f:
			json.dump(items, f, ensure_ascii=False, indent=2)

	agents_path = os.path.join(OUTPUT_ROOT, OUTPUT_FILE)
	agent_rows = []
	for record in records:
		agent_rows.append({
			"name": record.get("agent_name", ""),
			"description": record.get("agent_description", ""),
			"icon": record.get("agent_icon", ""),
			"tags": record.get("agent_tags", []),
			"prompt": record.get("agent_prompt", ""),
			"source_url": record.get("url", ""),
			"row_no": record.get("no", ""),
		})
	with open(agents_path, "w", encoding="utf-8") as f:
		json.dump(agent_rows, f, ensure_ascii=False, indent=2)

	index_path = os.path.join(OUTPUT_ROOT, "index.json")
	with open(index_path, "w", encoding="utf-8") as f:
		json.dump(sorted(categorized.keys()), f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
	main()
